from __future__ import annotations

import os
from typing import cast

import openpyxl  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

wb = openpyxl.Workbook()
ws: Worksheet = cast(Worksheet, wb.active)
ws.title = "Reorganization Plan"

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")
section_font = Font(name="Arial", bold=True, size=11)
section_fill = PatternFill("solid", fgColor="D6E4F0")
normal_font = Font(name="Arial", size=10)
italic_font = Font(name="Arial", size=10, italic=True, color="666666")
question_fill = PatternFill("solid", fgColor="FFF2CC")
thin_border = Border(bottom=Side(style="thin", color="D0D0D0"))

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 50
ws.column_dimensions["G"].width = 50

headers: list[str] = [
    "Depth",
    "Proposed Path",
    "Source (current location)",
    "Files",
    "Size (GB)",
    "Notes",
    "Your Comments / Changes",
]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 25
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:G1"

row: int = 2


def add_section(
    title: str, source: str = "", files: int | str = "", size: float | str = "", notes: str = ""
) -> None:
    global row
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = section_fill
    ws.cell(row=row, column=1, value=0).font = section_font
    ws.cell(row=row, column=2, value=title).font = section_font
    ws.cell(row=row, column=3, value=source).font = normal_font
    if files:
        ws.cell(row=row, column=4, value=files).font = normal_font
    if size:
        ws.cell(
            row=row, column=5, value=round(size, 2) if isinstance(size, float) else size
        ).font = normal_font
    ws.cell(row=row, column=6, value=notes).font = normal_font
    ws.row_dimensions[row].height = 22
    row += 1


def add_sub(
    depth: int,
    path: str,
    source: str = "",
    files: int | str = "",
    size: float | str = "",
    notes: str = "",
) -> None:
    global row
    indent = "    " * (depth - 1)
    ws.cell(row=row, column=1, value=depth).font = normal_font
    ws.cell(row=row, column=2, value=indent + path).font = normal_font
    ws.cell(row=row, column=3, value=source).font = normal_font
    if files:
        ws.cell(row=row, column=4, value=files).font = normal_font
    if size:
        ws.cell(
            row=row, column=5, value=round(size, 2) if isinstance(size, float) else size
        ).font = normal_font
    ws.cell(row=row, column=6, value=notes).font = italic_font
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = thin_border
    row += 1


def add_question(text: str) -> None:
    global row
    ws.cell(row=row, column=2, value=text).font = Font(
        name="Arial", size=10, italic=True, color="996600"
    )
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = question_fill
    row += 1


def add_blank() -> None:
    global row
    row += 1


# ============================================================
# WORK
# ============================================================
add_section("Work/", notes="All professional/employment work")

add_sub(1, "Coca-Cola/", "merge multiple sources", 10903, 11.7, "Primary source: Coke/Current/")
add_sub(2, "Admin/", "Coke/Current/Admin + Admin/ (top-level)", 266, 0.115)
add_sub(3, "Benefits/", "", 5)
add_sub(3, "HR/", "", 37, notes="Comp statements, interviewing")
add_sub(3, "Interview Guides/", "", 3)
add_sub(3, "Status/", "", 70, notes="Weekly status .docx files")
add_sub(3, "Time and Exp/", "", 144, 0.104, "Expense reports + trip receipts")
add_sub(2, "Business Services/", "Coke/Current/Bus Svcs + Business Services/ (top)", 469, 0.945)
add_sub(3, "TSAD/", "merge 3 TSAD dirs", 413, 0.85, "Largest sub-area")
add_sub(2, "BSNA/", "Coke/BSNA + BSNA Backup/", 406, 1.44)
add_sub(3, "BSNA Transformation/", "", 159, 0.594, "Includes McKinsey, Working Docs")
add_sub(3, "Bottler Comms/", "", 9, 0.111)
add_sub(3, "CONA/", "", 13, 0.049)
add_sub(3, "Exec Decks/", "", 15, 0.188)
add_sub(3, "Finance/", "", 23, 0.004)
add_sub(3, "Technology/", "", 45, 0.319)
add_sub(2, "Projects/", "Coke/Current/Projects", 955, 1.1)
add_sub(3, "BASIS Replacement/", "", 162, 0.184)
add_sub(3, "System of the Future/", "", 776, 0.926, "CONA/Deployment sub-tree")
add_sub(3, "Customer Lifecycle/", "", 17, 0.013)
add_sub(2, "Strategy and BTO/", "Coke/Current/Strategy and BTO", 204, 0.23)
add_sub(3, "GTM/", "", 9)
add_sub(3, "Integrated Process/", "", 155, 0.141, "Includes McKinsey sub-dir")
add_sub(3, "McKinsey ZIP/", "", 22, 0.037)
add_sub(2, "Commercial Leadership/", "Coke/Current/Commercial Leadership", 21, 1.6)
add_sub(3, "ISAM/", "", 4)
add_sub(3, "Small Store Segmentation/", "", 7, 1.5)
add_sub(3, "WCPPT/", "", 5)
add_sub(2, "Customer Care/", "Coke/Current/Customer Care", 33, 0.044)
add_sub(
    2,
    "Email Downloads/",
    "Coke/Current/Email Downloads",
    6802,
    4.6,
    "Keep as-is, bulk Outlook export",
)
add_sub(3, "2013-03-01/", "", 945, 2.1)
add_sub(3, "2013-08-28/", "", 5857, 2.5)
add_sub(2, "General Coke IT and Business/", "", 86, 0.375)
add_sub(2, "Region Sales/", "", 43, 0.218)
add_sub(2, "CokeONE and SCALE/", "", 19, 0.028)
add_sub(2, "National Retail Sales/", "", 5, 0.088)
add_sub(2, "Shared Services/", "", 6, 0.011)
add_sub(2, "Supply Chain/", "", 6, 0.017)
add_sub(2, "QA/", "", 4, 0.001)
add_sub(2, "Python/", "Coke/Python", 103, 0.244, "Data processing scripts (.upd/.dwn)")
add_sub(2, "Programming/", "Coke/Current/Programming", 27, 0.039, "Python scripts + test data")
add_question(
    "QUESTION: Documents/Coke (~7,924 files) mirrors Coke/Current. After dedup, merge unique remainders into this tree?"
)
add_blank()

add_sub(1, "Stratfield/", "Business/Stratfield", 1816, 3.3, "Active consulting practice")
add_sub(2, "Chick-Fil-A/", "", 1586, 2.4, "Biggest client")
add_sub(3, "CFA VIS/", "", 1230, 0.775)
add_sub(3, "DTT Scorecard/", "", 42, 0.248)
add_sub(3, "Support Now/", "", 109, 0.71)
add_sub(3, "SEMS/", "", 33, 0.078)
add_sub(3, "SC Innovation/", "", 40, 0.053)
add_sub(3, "Connected Workplace/", "", 13, 0.056)
add_sub(3, "EAFSM/", "", 22, 0.004)
add_sub(3, "CFA Now/", "", 15, 0.006)
add_sub(3, "Process Improvement/", "", 16, 0.045)
add_sub(3, "2024 Planning/", "", 2)
add_sub(2, "ATS-CRM/", "", 158, 0.561, "Job Diva + RCRM data exports")
add_sub(2, "AI/", "", 13, 0.174)
add_sub(2, "Acentra/", "", 9, 0.067)
add_sub(2, "Bus Dev/", "", 13, 0.017)
add_sub(2, "Invesco/", "", 1)
add_sub(2, "PGRX/", "", 4, 0.006)
add_sub(2, "Practice Dev/", "", 6, 0.003)
add_sub(2, "QA/", "", 6, 0.002)
add_sub(2, "JDs/", "", 3)
add_sub(2, "Internal/", "", 5, notes="Receipts sub-dir")
add_blank()

add_sub(1, "Consulting/", "Consulting/ (top-level)", 123, 0.178, "Pre-Stratfield consulting")
add_sub(2, "Chick-Fil-A/", "", 39, 0.136, "Earlier CFA work")
add_sub(2, "Quest Renewables/", "", 19, 0.004)
add_sub(2, "Safely/", "", 24, 0.024)
add_sub(2, "Rubik/", "", 1)
add_sub(2, "Admin/", "Consulting/Stratfield/Admin", 36, 0.013, "Receipts")
add_question(
    "QUESTION: Merge Consulting/Chick-Fil-A into Stratfield/Chick-Fil-A? Or keep separate (different era)?"
)
add_blank()

add_sub(1, "Sellr/", "Sellr/", 30, 0.078)
add_sub(2, "Admin/Employment/", "", 15, 0.042)
add_sub(2, "Finance/", "", 3, 0.003)
add_sub(2, "Marketing/", "", 5, 0.001)
add_blank()

add_sub(1, "Ginkgo/", "Business/Ginkgo", 184, 0.067)
add_sub(2, "Graphics/", "", 145, 0.034)
add_sub(2, "The Ginkgo Link/", "", 29, 0.029, "Brand guide, logo, social media")
add_sub(2, "Blog Posts/", "", 7, 0.003)
add_blank()

add_sub(1, "GV/", "GV Documents/", 2, 0, "2 pptx files")
add_blank()

# ============================================================
# AMATEUR RADIO
# ============================================================
add_section("Amateur Radio/", "keep as-is", 6961, 11.0, "Well-organized, minimal changes")
add_sub(1, "ARES/", "", 160, 0.181)
add_sub(1, "AUXCOM/", "", 58, 0.083)
add_sub(1, "Antennas/", "", 42, 0.116)
add_sub(1, "CW - Morse Code/", "", 461, 0.012, "K7QO Course, keys, RUFZXP")
add_sub(1, "Contests/", "", 18, 0.021)
add_sub(1, "Equipment/", "", 211, 0.259, "Anytone, Elecraft, Kenwood, Quansheng, Yaesu")
add_sub(1, "Licenses/", "", 21, 0.006)
add_sub(1, "Logging/", "", 13, 0.001)
add_sub(1, "MARS/", "", 1747, 4.2, "Largest sub-dir (incl Software Installed)")
add_sub(1, "Manuals/", "", 70, 0.311)
add_sub(
    1,
    "N3FJP Software/",
    "Documents/Affirmatech/N3FJP Software",
    368,
    0.033,
    "Contest logging - move from Documents",
)
add_sub(1, "Newsletters/", "", 3, 0.032)
add_sub(1, "POTA SOTA/", "", 5, 0.016)
add_sub(1, "Projects/", "", 9, 0.096, "NanoKeyer, NorCal 40a, Raspberry Pi")
add_sub(1, "QRP/", "", 194, 2.9, "193 QQ PDFs")
add_sub(1, "QSLs/", "", 38, 0.059)
add_sub(1, "Reference Material/", "", 2159, 1.9, "ARRL Handbook + Antenna Book")
add_sub(1, "SHARES/", "", 15, 0.011)
add_sub(1, "Silver Comet/", "", 1126, 0.412, "Club web site backup")
add_sub(1, "Software/", "", 549, 0.17)
add_sub(1, "TQSL/", "", 5)
add_sub(1, "Go-Box/", "Projects/Amateur Radio/Go-Box", 8, 0.057, "Move from Projects/")
add_blank()

# ============================================================
# SAILING
# ============================================================
add_section("Sailing/", "merge Sailing/ + Boat/", 152, 0.703)
add_sub(1, "Events/", "Sailing/ year dirs", 100, 0.262, "2018-2026 regattas")
add_sub(1, "Boat Manuals/", "Boat/", 15, 0.346, "15 PDFs")
add_sub(1, "Charts/", "Documents/Charts", 1002, 0.773)
add_sub(2, "ENC/US_REGION02/", "", 692, 0.177)
add_sub(2, "RNC/US_REGION02/", "", 310, 0.596)
add_sub(1, "RC Barge/", "", 3, 0.012)
add_sub(1, "Templates and Diagrams/", "", 2, 0.001)
add_sub(1, "US Sailing Docs/", "", 7, 0.009)
add_question("QUESTION: Charts under Sailing, or separate top-level Reference/Charts?")
add_blank()

# ============================================================
# MAKING
# ============================================================
add_section("Making/", notes="All hands-on/shop/build hobbies")

add_sub(1, "3D Printing/", "merge 3D Printer Stuff + Projects/3D Printing", 1828, 5.84)
add_sub(2, "Designs/", "Projects/3D Printing/designs", 784, 0.555, "40+ project sub-dirs")
add_sub(2, "Gcode/", "Projects/3D Printing/gcode", 285, 3.7)
add_sub(2, "Printer Mods/", "3D Printer Stuff/AM8*, Anet*, Prusa*", 400, 0.1)
add_sub(2, "Camera Stuff/", "", 24, 0.184)
add_sub(2, "Covid Shields/", "", 12, 0.253)
add_sub(2, "Gridfinity/", "", 19, 0.004)
add_sub(2, "For Troy to Look At/", "", 15, 0.006)
add_blank()

add_sub(1, "Woodworking/", "Workshop/", 2188, 9.34)
add_sub(2, "Fine Woodworking/", "Workshop/FWW Issues", 244, 3.0, "244 issue PDFs")
add_sub(2, "American Woodworker/", "Workshop/T4570_...Magazine", 191, 3.1, "By year 1985-2014")
add_sub(2, "1000 Tips and Tricks/", "", 1678, 0.114, "plansnow + woodsmithstore catalogs")
add_sub(2, "Bandsaw Plans/", "Personal/bandsaw_plans", 173, 0.006)
add_sub(2, "Lawnchair Plans/", "Personal/lawnchair", 116, 0.004)
add_blank()

add_sub(1, "CNC/", "Projects/CNC Milling", 49, 0.417)
add_sub(2, "Reference Material/", "", 30, 0.407)
add_sub(2, "Designs/", "", 4)
add_sub(2, "Gcode/", "", 14, 0.01)
add_sub(2, "Joe's CNC Model 2006/", "Personal/Joes CNC Model 2006 R-2", 53, 0.028)
add_blank()

add_sub(1, "Electronics/", "merge Projects/Electronics + PCB + Electronics/", 205, 0.193)
add_sub(2, "KX3 Interface/", "", 57, 0.013)
add_sub(2, "KX3 Audio Switcher/", "Projects/PCB/", 46, 0.007)
add_sub(2, "Digirig/", "", 66, 0.005)
add_sub(2, "KiCAD Like a Pro/", "", 9, 0.064)
add_sub(2, "Adafruit/", "", 10, 0.004)
add_sub(2, "Manuals/", "Electronics/ (top-level)", 2, 0.066)
add_blank()

# ============================================================
# PERSONAL
# ============================================================
add_section("Personal/", "Personal/", 2309, 3.5)

add_sub(1, "Career/", "Personal/Career", 54, 0.009, "Resume, job apps, offers")
add_sub(2, "Troy Davis Background/", "", 13, notes="All versions - protected from dedup")
add_sub(2, "CGI/", "", 6, 0.003)
add_sub(2, "Croixstone/", "", 2)
add_sub(2, "Experian/", "", 3)
add_sub(2, "GA AOC/", "", 4, 0.001)
add_sub(2, "Harrison - Physics/", "", 3)
add_sub(2, "Kennesaw/", "", 2)
add_sub(2, "Tosca/", "", 1)
add_blank()

add_sub(1, "Family/", "Personal/Family", 418, 0.41)
add_sub(2, "Finance/", "", 291, 0.278, "Quicken backups, credit, TurboTax")
add_sub(2, "Genealogy/", "", 102, 0.066, "By surname")
add_sub(2, "Daniel/", "Personal/Daniel", 16, 0.033)
add_sub(2, "Jamie/", "Personal/Jamie", 14, 0.011)
add_blank()

add_sub(1, "Finance/", "Personal/Finance", 43, 0.024)
add_sub(2, "Insurance/", "Personal/Insurance", 4, 0.001)
add_sub(2, "Taxes/", "Personal/Taxes", 43, 0.052, "By year: 2020-2024")
add_sub(2, "Davis Venture Investments/", "", 1)
add_blank()

add_sub(1, "Health/", "Personal/Health + Health/ (top-level)", 21, 0.028)
add_blank()

add_sub(1, "Home/", "")
add_sub(2, "3381 Valley Hill/", "Personal/3381* + mortgage docs", 29, 0.083)
add_sub(2, "Cabin/", "Personal/Cabin", 4, 0.009)
add_sub(2, "Real Estate/", "Real Estate/ (top-level)", 1, 0.01)
add_blank()

add_sub(1, "Travel/", "Trips/ + Personal trip dirs", 20, 0.118)
add_sub(2, "Newport Trip/", "", 15, 0.085)
add_sub(2, "2023 Italy-France/", "", 2, 0.033)
add_sub(2, "BVI 2013/", "root-level BVI files", 3)
add_sub(2, "Maps/", "Maps-Trips/", 5)
add_blank()

add_sub(1, "Education/", "Education/ + Personal/KSU", 11)
add_sub(2, "KSU/", "", 6, 0.011)
add_sub(2, "Bible Study/", "Personal/Bible Study", 19, 0.037)
add_blank()

add_sub(1, "Books/", "")
add_sub(2, "Kindle Backup/", "Personal/Kindle Backup", 456, 0.348)
add_sub(2, "Guitar/", "Personal/Guitar/Beato", 64, 0.044, "Rick Beato PDFs")
add_blank()

add_sub(1, "Vehicles/", "Vehicles/ + Jeep/ + Personal/Vehicles", 13, 0.09)
add_sub(1, "Writing/", "Personal/Writing", 30, 0.001, "Energy.scriv")
add_sub(1, "Scouts/", "", 1)
add_sub(1, "First Lego League/", "First Lego League/", 69, 0.05)
add_sub(1, "Moms 80th/", "", 2)
add_sub(1, "Tech/", "Personal/Tech", 290, 1.6, "Fonts, Winget backups")
add_sub(1, "Jimmys Campaign 2018/", "", 2)
add_blank()

# ============================================================
# PROJECTS
# ============================================================
add_section("Projects/", notes="Dev projects, code, experiments")

add_sub(1, "Code/", "Projects/Code", 794, 5.2)
add_sub(2, "Python/", "", 707, 5.1)
add_sub(3, "practicalnlp/", "", 345, 0.16)
add_sub(3, "islp/", "", 80, 0.081)
add_sub(3, "stratfield/", "", 204, 4.9, "Client data files")
add_sub(3, "pdf_mimic/", "", 45)
add_sub(2, "AI/", "", 22, 0.026, "Agent-Memory project")
add_sub(2, "Vibe Coding Prompts/", "", 14)
add_sub(2, "saas_apps/immiflow/", "", 12)
add_question(
    "QUESTION: Projects/Code/Python/stratfield (4.9 GB data) - keep here or move to Work/Stratfield/Data?"
)
add_blank()

add_sub(1, "AI Projects/", "Projects/AI", 55, 0.006)
add_sub(2, "emcom-forms/", "", 47, 0.004)
add_sub(2, "CS-R OB24 Contract Review/", "", 7, 0.002)
add_blank()

add_sub(1, "Raspberry Pi/", "Projects/Raspberry Pi", 163, 0.029)
add_blank()

# ============================================================
# REFERENCE
# ============================================================
add_section("Reference/", notes="Non-project reference material")
add_sub(1, "Zoom Recordings/", "Documents/Zoom", 144, 1.8)
add_sub(1, "Office Templates/", "Documents/Templates", 5, 0.012)
add_sub(1, "Scans/", "Scanbot/ + Office Lens/ + OfficeMobile/", 22, 0.033)
add_sub(1, "ARIS Express/", "", 8, 0.002)
add_sub(1, "External Docs/", "", 5, 0.013)
add_sub(1, "iPad PDFs/", "Documents/iPad", 11, 0.033)
add_question("QUESTION: Zoom recordings (1.8 GB) - some are ham radio nets. Split?")
add_blank()

# ============================================================
# APP DATA
# ============================================================
add_section("App Data/", notes="Application configs synced via OneDrive")
add_sub(1, "KiCad/", "Documents/KiCad", 2139, 0.616, "v6 + v7 libraries")
add_sub(1, "PowerShell/", "Documents/PowerShell + WindowsPowerShell", 1011, 0.664)
add_sub(1, "N1MM Logger+/", "", 42, 0.012)
add_sub(1, "UniGetUI/", "", 63, 0.004)
add_sub(1, "EasyEDA-Pro/", "", 60, 0.055)
add_sub(1, "Kenwood/", "Documents/Kenwood", 6, notes="Radio configs")
add_sub(1, "Yaesu/", "Documents/Yaesu", 1)
add_sub(1, "Workspace/", "Documents/Workspace", 2628, 0.184, "Eclipse .metadata")
add_sub(1, "Claude/", "Documents/Claude", 4, notes="Scheduled tasks")
add_sub(1, "Cline/", "Documents/Cline", 10)
add_sub(1, "Kutools for Excel/", "", 48, 0.008)
add_sub(1, "OneNote Notebooks/", "", 4)
add_sub(1, "My Labels/", "", 3)
add_sub(1, "My Garmin/", "", 8, notes="Waypoint symbols")
add_sub(1, "Packet Engine Pro/", "", 3)
add_sub(1, "PGP/", "", 3, notes="Key rings")
add_sub(1, "OnScreen Control/", "", 3)
add_sub(1, "Fiddler2/", "", 3)
add_sub(1, "XSplit/", "", 7)
add_sub(1, "My Tableau Repository/", "", 7)
add_sub(1, "My Data Sources/", "", 6)
add_sub(1, "navigator/", "", 10, 0.065)
add_question("QUESTION: Keep App Data/ as top-level? Or nest under Reference/?")
add_blank()

# ============================================================
# ARCHIVE
# ============================================================
add_section("_Archive/", notes="Kept but not actively used. NOT deleted.")

add_sub(1, "SkyDrive/", "SkyDrive/", 6604, 6.09, "Old pre-OneDrive mirror (most removed by dedup)")
add_question(
    "QUESTION: After dedup, merge SkyDrive remnants into main structure? Or keep as _Archive/SkyDrive/?"
)
add_blank()

add_sub(1, "Favorites/", "all Favorites dirs", 696, 0, "All .url bookmark files")
add_sub(1, "Old Software/", "", notes="Installers and drivers")
add_sub(2, "VARA FM/", "", 1, 0.007)
add_sub(2, "VARA HF/", "", 1, 0.004)
add_sub(2, "VBCABLE Driver/", "", 27, 0.003)
add_sub(2, "WinTAK/", "", 2, 0.042)
add_sub(2, "TeraCopyPortable/", "", 12)
add_sub(2, "TaskSeparator11/", "", 42, 0.003)
add_sub(2, "Blueprint RC 2010/", "", 115, 0.03)
add_blank()

add_sub(1, "Root Files/", "", notes="~30 loose files from OneDrive root")
add_sub(2, "34 loose files", "", 34, notes="Need manual sort or keep here")
add_blank()

add_sub(1, "Misc/", "", notes="App artifacts and old data")
add_sub(2, "Visual Studio 2015/", "", 1)
add_sub(2, "CommunityPlugins/", "", 12, 0.011)
add_sub(2, "My Meetings/", "", 130, 0.001)
add_sub(2, "dumps/", "", 1, 0.027, "SQL dump")
add_sub(2, "git_tutorial/", "", 121, 0.001)
add_sub(2, "Dell SupportAssist/", "", 3)
add_sub(2, "HPrintJobsStorage/", "", 1)
add_sub(2, "G4FON Morse Trainer/", "", 1)
add_sub(2, "HDSDR/", "", 2)
add_sub(2, "EchoLink/", "", 4)
add_sub(2, "ExportBlock CSV/", "", 2)
add_sub(2, "Product Launch Workspace/", "", 2)
add_sub(2, "Home Share/", "", 6)
add_sub(2, "Apps/", "", 5)
add_sub(2, "Videos/", "", 2)
add_sub(2, "Public/", "", 5, 0.04)
add_sub(2, "Email attachments/", "", 16, 0.36)

# ============================================================
# SUMMARY SHEET
# ============================================================
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 15
ws2.column_dimensions["D"].width = 55

for col, h in enumerate(["Top-Level Directory", "Est. Files", "Est. Size (GB)", "Description"], 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
ws2.freeze_panes = "A2"

data: list[tuple[str, int, float, str]] = [
    ("Work/", 13000, 15.3, "Coca-Cola, Stratfield, Consulting, Sellr, Ginkgo, GV"),
    ("Amateur Radio/", 7330, 11.0, "Keep as-is + N3FJP from Documents"),
    ("Sailing/", 1170, 1.5, "Regattas, boat manuals, charts"),
    ("Making/", 4270, 15.8, "3D Printing, Woodworking, CNC, Electronics"),
    ("Personal/", 2600, 3.8, "Career, Family, Finance, Health, Home, Travel, Books"),
    ("Projects/", 1010, 5.2, "Code, AI, Raspberry Pi"),
    ("Reference/", 190, 1.9, "Zoom recordings, templates, scans"),
    ("App Data/", 4010, 1.6, "KiCad, PowerShell, app configs from Documents/"),
    ("_Archive/", 7700, 6.6, "SkyDrive mirror, old bookmarks, installers, misc"),
]

for i, (name, files, size, desc) in enumerate(data, 2):
    ws2.cell(row=i, column=1, value=name).font = Font(name="Arial", bold=True, size=10)
    c = ws2.cell(row=i, column=2, value=files)
    c.number_format = "#,##0"
    c = ws2.cell(row=i, column=3, value=size)
    c.number_format = "0.0"
    ws2.cell(row=i, column=4, value=desc).font = Font(name="Arial", size=10)

r = len(data) + 2
ws2.cell(row=r, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=r, column=2, value=f"=SUM(B2:B{r-1})").number_format = "#,##0"
ws2.cell(row=r, column=3, value=f"=SUM(C2:C{r-1})").number_format = "0.0"

out: str = os.path.join(
    os.environ.get("USERPROFILE", ""), "OneDrive", "Desktop", "OneDrive Reorganization Plan.xlsx"
)
wb.save(out)
print(f"Saved to {out}")
